#!/usr/bin/env python3
"""
Import script to migrate data from DELTA 2026 Excel file to Supabase.
Reads finalized and in-progress dossiers from consultant sheets and imports them.
"""

import os
import sys
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple
from decimal import Decimal

import openpyxl
from openpyxl.utils import get_column_letter
from supabase import create_client, Client


# Configuration
EXCEL_FILE = "/sessions/wonderful-zen-hypatia/mnt/.projects/019d297c-190d-7091-b8c7-3e10bda319c2/files/019d297f-5965-77c1-afcc-29ba7a82d885.xlsx"

CONSULTANTS_SHEETS = {
    "Stéphane - Finalisée": "Stéphane",
    "Stéphane - En cours": "Stéphane",
    "POOL - Finalisée": "POOL",
    "POOL - En cours": "POOL",
    "Maxine - Finalisée": "Maxine",
    "Maxine - En cours": "Maxine",
    "Thélo - Finalisée": "Thélo",
    "Thélo - En cours": "Thélo",
    "Mathias - Finalisée": "Mathias",
    "Mathias - En cours": "Mathias",
    "Guillaume - Finalisée": "Guillaume",
    "Guillaume - En cours": "Guillaume",
    "James - Finalisée": "James",
    "James - En cours": "James",
    "Hugues - Finalisée": "Hugues",
    "Hugues - En cours": "Hugues",
    "Gilles - Finalisée": "Gilles",
    "Gilles - En cours": "Gilles",
    "Valentin - Finalisée": "Valentin",
    "Valentin - En cours": "Valentin",
    "Sylvain - Finalisée": "Sylvain",
}


class DeltaImporter:
    def __init__(self):
        self.wb = None
        self.supabase: Optional[Client] = None
        self.consultants: Dict[str, str] = {}  # nom -> id
        self.produits: Dict[str, str] = {}  # nom -> id
        self.compagnies: Dict[str, str] = {}  # nom -> id
        self.clients_cache: Dict[str, str] = {}  # (nom, prenom) -> id

        self.stats = {
            "clients_created": 0,
            "dossiers_created": 0,
            "errors": [],
        }

    def connect_supabase(self) -> bool:
        """Connect to Supabase using environment variables."""
        url = os.getenv("SUPABASE_URL")
        key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

        if not url or not key:
            print("ERROR: Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")
            return False

        self.supabase = create_client(url, key)
        print(f"Connected to Supabase: {url}")
        return True

    def load_excel(self) -> bool:
        """Load the Excel workbook."""
        if not Path(EXCEL_FILE).exists():
            print(f"ERROR: Excel file not found: {EXCEL_FILE}")
            return False

        self.wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        print(f"Loaded Excel file: {EXCEL_FILE}")
        print(f"Available sheets: {len(self.wb.sheetnames)}")
        return True

    def load_reference_data(self) -> bool:
        """Load consultants, produits, and compagnies from Supabase."""
        try:
            # Load consultants
            result = self.supabase.table("consultants").select("id, nom, prenom").execute()
            for row in result.data:
                key = row["nom"]
                self.consultants[key] = row["id"]
            print(f"Loaded {len(self.consultants)} consultants")

            # Load produits
            result = self.supabase.table("produits").select("id, nom").execute()
            for row in result.data:
                self.produits[row["nom"]] = row["id"]
            print(f"Loaded {len(self.produits)} produits")

            # Load compagnies
            result = self.supabase.table("compagnies").select("id, nom").execute()
            for row in result.data:
                self.compagnies[row["nom"]] = row["id"]
            print(f"Loaded {len(self.compagnies)} compagnies")

            return True
        except Exception as e:
            print(f"ERROR loading reference data: {e}")
            return False

    def parse_montant(self, value) -> float:
        """Parse montant value, handling various formats."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", "."))
            except:
                return 0.0
        return 0.0

    def parse_date(self, value) -> Optional[str]:
        """Parse date value, returning ISO format or None."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, str):
            try:
                # Try parsing common formats
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        dt = datetime.strptime(value, fmt)
                        return dt.date().isoformat()
                    except:
                        pass
                return None
            except:
                return None
        return None

    def normalize_compagnie(self, compagnie_str: str) -> str:
        """Normalize compagnie string to match database."""
        if not compagnie_str:
            return None

        compagnie_str = str(compagnie_str).strip()

        # Extract base compagnie name (before dash)
        if " - " in compagnie_str:
            compagnie_str = compagnie_str.split(" - ")[1].strip()

        return compagnie_str if compagnie_str else None

    def find_or_create_client(self, nom: str, prenom: str, pays: str = "FRANCE") -> Optional[str]:
        """Find or create a client, returning their ID."""
        if not nom:
            return None

        # Normalize inputs
        nom = str(nom).strip()
        prenom = str(prenom).strip() if prenom else ""
        pays = str(pays).strip() if pays else "FRANCE"

        cache_key = (nom, prenom)
        if cache_key in self.clients_cache:
            return self.clients_cache[cache_key]

        try:
            # Try to find existing client
            result = self.supabase.table("clients").select("id").where(
                "nom", "eq", nom
            ).where(
                "prenom", "eq", prenom
            ).execute()

            if result.data:
                client_id = result.data[0]["id"]
                self.clients_cache[cache_key] = client_id
                return client_id

            # Create new client
            result = self.supabase.table("clients").insert({
                "nom": nom,
                "prenom": prenom,
                "pays": pays,
                "statut_kyc": "non",
            }).execute()

            if result.data:
                client_id = result.data[0]["id"]
                self.clients_cache[cache_key] = client_id
                self.stats["clients_created"] += 1
                return client_id

            return None
        except Exception as e:
            error_msg = f"Error creating client {nom} {prenom}: {e}"
            self.stats["errors"].append(error_msg)
            print(f"  WARNING: {error_msg}")
            return None

    def find_or_create_produit(self, nom: str) -> Optional[str]:
        """Find or create a produit, returning their ID."""
        if not nom:
            return None

        nom = str(nom).strip()

        if nom in self.produits:
            return self.produits[nom]

        try:
            # Create new produit
            result = self.supabase.table("produits").insert({
                "nom": nom,
                "categorie": "autre",
            }).execute()

            if result.data:
                produit_id = result.data[0]["id"]
                self.produits[nom] = produit_id
                return produit_id

            return None
        except Exception as e:
            print(f"  WARNING: Error creating produit {nom}: {e}")
            return None

    def find_or_create_compagnie(self, nom: str) -> Optional[str]:
        """Find or create a compagnie, returning their ID."""
        if not nom:
            return None

        nom = str(nom).strip()

        if nom in self.compagnies:
            return self.compagnies[nom]

        try:
            # Create new compagnie
            result = self.supabase.table("compagnies").insert({
                "nom": nom,
                "taux_defaut": 0.0,
            }).execute()

            if result.data:
                compagnie_id = result.data[0]["id"]
                self.compagnies[nom] = compagnie_id
                return compagnie_id

            return None
        except Exception as e:
            print(f"  WARNING: Error creating compagnie {nom}: {e}")
            return None

    def parse_sheet(self, sheet_name: str, is_finalisee: bool) -> List[Dict]:
        """
        Parse a consultant sheet and return list of dossier data.
        Headers are expected at row 2: NOM, PRENOM, PRODUIT, COMPAGNIE, MONTANT, FINANCEMENT, COMMENTAIRE, DATE, PAYS
        """
        ws = self.wb[sheet_name]
        dossiers = []

        # Find column indices from row 2 headers
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(2, col_idx).value
            if header:
                headers[str(header).strip().upper()] = col_idx

        if not headers:
            print(f"  WARNING: No headers found in {sheet_name}")
            return dossiers

        # Map column names
        col_nom = headers.get("NOM")
        col_prenom = headers.get("PRENOM")
        col_produit = headers.get("PRODUIT")
        col_compagnie = headers.get("COMPAGNIE")
        col_montant = headers.get("MONTANT")
        col_financement = headers.get("FINANCEMENT")
        col_commentaire = headers.get("COMMENTAIRE")
        col_date = headers.get("DATE")
        col_pays = headers.get("PAYS")

        if not all([col_nom, col_prenom, col_montant]):
            print(f"  WARNING: Missing required columns in {sheet_name}")
            return dossiers

        # Parse data rows (starting from row 3)
        for row_idx in range(3, ws.max_row + 1):
            nom = ws.cell(row_idx, col_nom).value
            if not nom or not str(nom).strip():
                continue

            nom = str(nom).strip()
            prenom = ws.cell(row_idx, col_prenom).value
            prenom = str(prenom).strip() if prenom else ""

            montant = self.parse_montant(ws.cell(row_idx, col_montant).value)
            if montant == 0:
                continue

            dossier = {
                "nom": nom,
                "prenom": prenom,
                "montant": montant,
            }

            if col_produit:
                produit = ws.cell(row_idx, col_produit).value
                dossier["produit"] = str(produit).strip() if produit else None

            if col_compagnie:
                compagnie = ws.cell(row_idx, col_compagnie).value
                dossier["compagnie"] = self.normalize_compagnie(compagnie)

            if col_financement:
                financement = ws.cell(row_idx, col_financement).value
                dossier["financement"] = str(financement).strip().lower() if financement else "cash"
            else:
                dossier["financement"] = "cash"

            if col_commentaire:
                commentaire = ws.cell(row_idx, col_commentaire).value
                dossier["commentaire"] = str(commentaire).strip() if commentaire else None

            if col_date:
                date_val = ws.cell(row_idx, col_date).value
                dossier["date"] = self.parse_date(date_val)

            if col_pays:
                pays = ws.cell(row_idx, col_pays).value
                dossier["pays"] = str(pays).strip().upper() if pays else "FRANCE"
            else:
                dossier["pays"] = "FRANCE"

            dossier["is_finalisee"] = is_finalisee
            dossiers.append(dossier)

        return dossiers

    def import_dossier(self, sheet_name: str, consultant_name: str, dossier: Dict) -> bool:
        """Import a single dossier to Supabase."""
        try:
            # Get consultant ID
            consultant_id = self.consultants.get(consultant_name)
            if not consultant_id:
                self.stats["errors"].append(f"Consultant not found: {consultant_name}")
                return False

            # Find or create client
            client_id = self.find_or_create_client(
                dossier["nom"],
                dossier["prenom"],
                dossier.get("pays", "FRANCE")
            )
            if not client_id:
                self.stats["errors"].append(f"Failed to create client: {dossier['nom']}")
                return False

            # Find or create produit
            produit_id = None
            if dossier.get("produit"):
                produit_id = self.find_or_create_produit(dossier["produit"])

            # Find or create compagnie
            compagnie_id = None
            if dossier.get("compagnie"):
                compagnie_id = self.find_or_create_compagnie(dossier["compagnie"])

            # Normalize financement
            financement = dossier.get("financement", "cash").lower()
            if financement not in ["cash", "credit", "lombard", "remploi"]:
                financement = "cash"

            # Determine statut
            statut = "client_finalise" if dossier["is_finalisee"] else "client_en_cours"

            # Insert dossier
            result = self.supabase.table("dossiers").insert({
                "client_id": client_id,
                "consultant_id": consultant_id,
                "produit_id": produit_id,
                "compagnie_id": compagnie_id,
                "montant": dossier["montant"],
                "financement": financement,
                "statut": statut,
                "commentaire": dossier.get("commentaire"),
                "date_operation": dossier.get("date") or datetime.now().date().isoformat(),
            }).execute()

            if result.data:
                self.stats["dossiers_created"] += 1
                return True

            return False
        except Exception as e:
            error_msg = f"Error importing dossier for {dossier['nom']}: {e}"
            self.stats["errors"].append(error_msg)
            print(f"  WARNING: {error_msg}")
            return False

    def run(self) -> bool:
        """Run the complete import process."""
        print("\n=== DELTA 2026 Excel to Supabase Import ===\n")

        # Connect to Supabase
        if not self.connect_supabase():
            return False

        # Load Excel
        if not self.load_excel():
            return False

        # Load reference data
        if not self.load_reference_data():
            return False

        print("\n=== Processing Sheets ===\n")

        # Process each consultant sheet
        for sheet_name, consultant_name in CONSULTANTS_SHEETS.items():
            if sheet_name not in self.wb.sheetnames:
                print(f"SKIP: Sheet not found: {sheet_name}")
                continue

            is_finalisee = "Finalisée" in sheet_name
            status_label = "Finalisée" if is_finalisee else "En cours"

            print(f"Processing {sheet_name}...")

            dossiers = self.parse_sheet(sheet_name, is_finalisee)
            print(f"  Found {len(dossiers)} dossiers")

            for dossier in dossiers:
                self.import_dossier(sheet_name, consultant_name, dossier)

        # Print summary
        self.print_summary()
        return True

    def print_summary(self):
        """Print import summary."""
        print("\n=== Import Summary ===\n")
        print(f"Clients created: {self.stats['clients_created']}")
        print(f"Dossiers created: {self.stats['dossiers_created']}")
        print(f"Errors: {len(self.stats['errors'])}")

        if self.stats["errors"]:
            print("\nErrors encountered:")
            for error in self.stats["errors"][:10]:
                print(f"  - {error}")
            if len(self.stats["errors"]) > 10:
                print(f"  ... and {len(self.stats['errors']) - 10} more errors")


def main():
    """Main entry point."""
    importer = DeltaImporter()
    success = importer.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
